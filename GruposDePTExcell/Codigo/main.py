import sys
import os

#from GruposDePTExcell.app import archivoGlobal1

import openpyxl 
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection


#archivoGlobal1.strip()


### variables globales 
grupos = []

# Define la ruta base donde se encuentra el proyecto
base_dir = os.path.expanduser('C:\\Users\\serle\\Desktop\\PROYECTO\\utilTech')
# Genera la ruta hasta archivosSubidos
directorio_archiGruposDePtExcell = os.path.join(base_dir, 'GruposDePTExcell', 'archivosSubidos\\')
# Genera la ruta hasta archivosGenerados
directorio_archiGeneradosGruposDePtExcell = os.path.join(base_dir, 'GruposDePTExcell', 'ArchivosGenerados\\')
#directorio_archiGruposDePtExcell = directorio_automatico + "\\"


directorio_automatico = os.path.dirname(os.path.abspath(__file__))
directorio_ruta = directorio_automatico + '\\'
plantilla_excell = 'PlantillaAsignacionCoeficientesPT.xlsx'
trimestre = '24-O'
excell_resultado = 'PlantillaAsignacionCoeficientesPT'+'-'+trimestre+'.xlsx' 
### abriendo el workbook
workbook= openpyxl.load_workbook(directorio_ruta+plantilla_excell)


### función para recibir el numero economico de un profesor
### TODO: falta poner condicion para cuando no ecnutra un profe, que regrese el valor 'N/A'
def get_numero_economico(nombre_del_profe):
    ### hoja de los profesores
    hoja_profesores = workbook['Profesores']

    ### fila y columna del profesor
    fila_encontrada = None
    columna_encontrada = None
    
    # Recorrer la hoja 'Profesores'
    for fila_index, fila in enumerate(hoja_profesores.iter_rows(values_only=True)):
        for columna_index, valor in enumerate(fila):
            
            if valor == None:
                continue
            elif ( (type(valor) == str) and nombre_del_profe in valor):
                # Almacena la posición del valor encontrado
                fila_encontrada = fila_index + 1  # Se suma 1 porque el índice comienza en 0
                columna_encontrada = columna_index + 1  # Se suma 1 porque el índice comienza en 0
                break  # Rompe el bucle interior si se encuentra el nombre
    
    if (fila_encontrada == None or columna_encontrada == None): return 'No se ha encontrado'
    
    else:
        numero_economino_celda = hoja_profesores.cell(row=fila_encontrada, column=columna_encontrada+1).value

        # print(fila_encontrada, columna_encontrada)
        # print(numero_economino_celda)

        return int(numero_economino_celda)


### función para manejar nombres por si alguien no tiene segundo nombre quitarle el 'SIN' para imprimir el nombre
def manejo_de_nombre(nombre):
    # print(nombre)

    ### si se trata de profesores
    if '.' in nombre:
        nombre = nombre.split(' ', 1)[-1]

    else:
        pass
    ### separar la estructura del nombre
    nombre_temporal = nombre.split(' ')

    ### si el ultimo elemento es igual a SIN, se lo quitamos
    if(nombre_temporal[-1] == 'SIN'):
        nombre_temporal.pop(-1)

    ### juntando las secciones del nombre
    nombre_final = ' '.join(nombre_temporal)

    return nombre_final

### funcion para generar un txt 
# para enlistar los alumnos con numero de PTs y matricula
def generar_txt_estudiantes_PT():

    print('generacion del archivo estudiantes de PT')
    print(f'La ruta del directorio automatico es: {directorio_archiGeneradosGruposDePtExcell}')

    ### crear el txt
    file = open(directorio_archiGeneradosGruposDePtExcell+"AlumnosPT123.txt","w+", encoding="utf-8")
    grupos_de_PT = ['1', '2', '3']

    ### loop para agregar información a la txt
    for element in grupos_de_PT:
        file.write('\nPT'+ element+'\n')
        for grupo in grupos:
            if grupo['PT'] == element:
                for alumno in grupo['Alumno(s)']:
                    file.write( manejo_de_nombre(alumno['nombre']) +' '+ alumno['matrícula']+ '\n')
    
    ### guardando el archivo txt
    file.close()


### funcion para agregar grupos 
# con base en el siguietne formato:
# {
#   'Asesor(es)': {'nombre': '', 'genero': ''} | [{'nombre': '', 'genero': ''},{'nombre': '', 'genero': ''}], 
#   'Alumno(s)': [{'nombre' : '', 'Matrícula': ''},{'nombre' : '', 'Matrícula': ''}] , 
#   'PT': '1' | '2' | '3'
# }
def manejo_grupo(proyecto : dict):
    
    # Agrega la información del proyecto actual a la lista de grupos
    
    grupo = {
        'Asesor(es)': '', 
        'Alumno(s)': [] , 
        'PT': ''
        }

    ### revisar si la lista de grupos está vacío
    if (len(grupos) == 0):
        # print('se cumple')
        grupo['Asesor(es)'] = proyecto['Asesor(es)']
        grupo['Alumno(s)'].append( {'nombre' : proyecto['Alumno'], 'matrícula': proyecto['Matrícula']} )
        grupo['PT'] = proyecto['PT']
        grupos.append(grupo)

    else:
        
        # Verificando si hay algún elemento en 'grupos' con 'Asesor(es)' y 'PT' iguales a los del proyecto
        indice_coincidencia = next((indice for indice, grupo in enumerate(grupos) if grupo['Asesor(es)'] == proyecto['Asesor(es)'] and grupo['PT'] == proyecto['PT']), None)


        if indice_coincidencia is not None:
            ### agregando el alumno al mismo grupo de PT
            # print("iguales")
            grupo_coincidente = grupos[indice_coincidencia]
            grupo_coincidente['Alumno(s)'].append({'nombre' : proyecto['Alumno'], 'matrícula': proyecto['Matrícula']})

        else:
            # print("no iguales")
            grupo['Asesor(es)'] = proyecto['Asesor(es)']
            grupo['Alumno(s)'].append( {'nombre' : proyecto['Alumno'], 'matrícula': proyecto['Matrícula']} )
            grupo['PT'] = proyecto['PT']
            grupos.append(grupo)


def crear_tabla(row_number, column_number, hoja, info):

    ### estilos globales
    alignment_header_tabla = Alignment(horizontal='center', vertical='center')
    font_header_tabla = Font(name='Arial', size=8, bold=True, italic=False, color='000000')
    font_header_info = Font(name='Arial', size=7, color='000000')
    font_tabla_info = Font(name='Arial', size=11, color='000000')
    border_header_tabla = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    
    horario_del_curso='14-17'


    # Nombre del curso
    hoja.merge_cells(start_row=row_number, start_column=column_number, end_row=(row_number+1), end_column=column_number)
    celda = hoja.cell(row=row_number, column=column_number)


    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'NOMBRE DEL CURSO'

    
    # Clave
    hoja.merge_cells(start_row=row_number, start_column=(column_number+1), end_row=(row_number+1), end_column=(column_number+1))
    celda = hoja.cell(row=row_number, column=column_number+1)

    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'CLAVE'


    # Grupo
    hoja.merge_cells(start_row=row_number, start_column=(column_number+2), end_row=(row_number+1), end_column=(column_number+2))
    celda = hoja.cell(row=row_number, column=column_number+2)


    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'GRUPO'


    # Cupo 
    celda = hoja.cell(row=row_number, column=column_number+3)

    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'CUPO'

    # Max 
    celda = hoja.cell(row=row_number+1, column=column_number+3)

    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'MAX.'


    # Profesor
    hoja.merge_cells(start_row=row_number, start_column=(column_number+4), end_row=(row_number+1), end_column=(column_number+4))
    celda = hoja.cell(row=row_number, column=column_number+4)
    
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'PROFESOR'


    # No. eco
    hoja.merge_cells(start_row=row_number, start_column=(column_number+5), end_row=(row_number+1), end_column=(column_number+5))

    celda = hoja.cell(row=row_number, column=column_number+5)
    
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'No. ECON.'


    # Horario y Aula
    hoja.merge_cells(start_row=row_number, start_column=column_number+6, end_row=(row_number), end_column=column_number+10)
    hoja.cell(row=row_number, column=(column_number+6), value='HORARIO  Y  AULA')
    celda = hoja.cell(row=row_number, column=(column_number+6))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    hoja.cell(row=row_number, column=(column_number+10)).border = border_header_tabla

    # LUNES, MARTES, MIERCOLES, JUEVES, VIERNES
    hoja.cell(row=row_number+1, column=(column_number+6), value='LUNES')
    celda = hoja.cell(row=row_number+1, column=(column_number+6))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla

    hoja.cell(row=row_number+1, column=(column_number+7), value='MARTES')
    celda = hoja.cell(row=row_number+1, column=(column_number+7))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla

    hoja.cell(row=row_number+1, column=(column_number+8), value='MIÉRCOLES')
    celda = hoja.cell(row=row_number+1, column=(column_number+8))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla


    hoja.cell(row=row_number+1, column=(column_number+9), value='JUEVES')
    celda = hoja.cell(row=row_number+1, column=(column_number+9))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla

    hoja.cell(row=row_number+1, column=(column_number+10), value='VIERNES')
    celda = hoja.cell(row=row_number+1, column=(column_number+10))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla



    if info['PT'] == '1': 
        pt = 'I' 
        horas_por_UEA = 9
    elif info['PT'] == '2': 
        pt = 'II'
        horas_por_UEA = 10
    else: 
        pt = 'III'
        horas_por_UEA = 10

    ### tabla observaciones
    if (info['Asesor(es)'] != 'POR ASIGNAR'):

        hoja.cell(row=row_number+1, column=(column_number+12), value='OBSERVACIONES')
        celda = hoja.cell(row=row_number+1, column=(column_number+12))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        hoja.cell(row=row_number+1, column=(column_number+13), value='Horas por alumno')
        celda = hoja.cell(row=row_number+1, column=(column_number+13))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla
        
        hoja.cell(row=row_number+1, column=(column_number+14), value='Alumnos')
        celda = hoja.cell(row=row_number+1, column=(column_number+14))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla
        
        hoja.cell(row=row_number+1, column=(column_number+15), value='Coeficiente por hr')
        celda = hoja.cell(row=row_number+1, column=(column_number+15))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        hoja.cell(row=row_number+1, column=(column_number+16), value='Coeficiente de participación')
        celda = hoja.cell(row=row_number+1, column=(column_number+16))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        hoja.cell(row=row_number+1, column=(column_number+17), value='Horas parciales')
        celda = hoja.cell(row=row_number+1, column=(column_number+17))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        hoja.cell(row=row_number+1, column=(column_number+18), value='Horas por UEA')
        celda = hoja.cell(row=row_number+1, column=(column_number+18))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        ### info llenado
        hoja.cell(row=row_number+2, column=(column_number+12), value='Asignación de profesor e inscripción de alumnos')
        celda = hoja.cell(row=row_number+2, column=(column_number+12))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        ### horas por alumnos valor
        hoja.cell(row=row_number+2, column=(column_number+13), value=2)
        celda = hoja.cell(row=row_number+2, column=(column_number+13))
        celda_horas_por_alumno = celda.value
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        ### Alumnos valor
        cantidad_de_alumnos = len(info['Alumno(s)']) 

        hoja.cell(row=row_number+2, column=(column_number+14), value=cantidad_de_alumnos)
        celda = hoja.cell(row=row_number+2, column=(column_number+14))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla
        
        ### Coeficiente por hr
        ### valor = horas por alumno / horas por UEA
        hoja.cell(row=row_number+2, column=(column_number+15), value=( round(celda_horas_por_alumno / horas_por_UEA, 2)))
        celda = hoja.cell(row=row_number+2, column=(column_number+15))
        celda_coeficiente_por_hora = celda.value
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla


        ### Coeficiente de part
        hoja.cell(row=row_number+2, column=(column_number+16), value= cantidad_de_alumnos* celda_coeficiente_por_hora)
        celda = hoja.cell(row=row_number+2, column=(column_number+16))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        ### Horas parciales
        hoja.cell(row=row_number+2, column=(column_number+17), value= cantidad_de_alumnos* 2)
        celda = hoja.cell(row=row_number+2, column=(column_number+17))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla


        ### Horas por UEA
        hoja.cell(row=row_number+2, column=(column_number+18), value=horas_por_UEA)
        celda = hoja.cell(row=row_number+2, column=(column_number+18))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        ### tabla nombre de los alumnos:
        ### DEJAR INSCRITOS A LOS SIGUIENTES ALUMNOS:
        celda = hoja.cell(row=row_number+4, column=column_number)
        celda.value = 'DEJAR INSCRITOS A LOS SIGUIENTES ALUMNOS:'
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla

        celda = hoja.cell(row=row_number+6, column=column_number)
        celda.value = 'Nombre'
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla

        celda = hoja.cell(row=row_number+6, column=column_number+1)
        celda.value = 'Matricula'
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla

        for index, alumno in enumerate(info['Alumno(s)']):
            celda_nombre = hoja.cell(row=row_number + 7 + index, column=column_number)
            celda_nombre.value = manejo_de_nombre(alumno['nombre'])

            celda_matricula = hoja.cell(row=row_number + 7 + index, column=column_number + 1)
            celda_matricula.value = alumno['matrícula']


    ### Llenando info
    # PROYECTO TERMINAL I
    celda = hoja.cell(row=row_number+2, column=column_number)


    celda.alignment = alignment_header_tabla
    celda.font = font_header_info
    celda.border = border_header_tabla
    celda.value = 'PROYECTO TERMINAL '+ pt

    # print('info ->', info)
    
    # Variable clave
    celda = hoja.cell(row=row_number+2, column=column_number+1)

    celda.alignment = alignment_header_tabla
    celda.font = font_header_info
    celda.border = border_header_tabla
    celda.value = int(info['CLAVE'])


    # Variable Grupo
    celda = hoja.cell(row=row_number+2, column=column_number+2)

    celda.alignment = alignment_header_tabla
    celda.font = Font(name='Calibri', size=11, color='000000')
    celda.border = border_header_tabla
    celda.value = info['GRUPO']

    

    # Variable Cupo máx
    celda = hoja.cell(row=row_number+2, column=column_number+3)

    celda.alignment = alignment_header_tabla
    celda.font = Font(name='Calibri', size=11, color='000000')
    celda.border = border_header_tabla
    celda.value = 30

    
    # Variable Profesor
    celda = hoja.cell(row=row_number+2, column=column_number+4)

    celda.alignment = alignment_header_tabla
    celda.font = Font(name='Calibri', size=11, color='000000')
    celda.border = border_header_tabla

    if info['Asesor(es)'] == 'POR ASIGNAR': 
        asesores = 'POR ASIGNAR'
        celda.value = asesores
    elif  isinstance( info['Asesor(es)'], list):
        # print('Its a list of professors')
        asesores = []

        for professor in info['Asesor(es)']:
            asesores.append(professor['nombre'])

        ### agregando el primer profe
        # print(asesores)
        celda.value = asesores[0]
        asesores.pop(0)

        for i in range(len(asesores)):
            y = i +1
            # print('valor de y -> ', y)
            # print('asesores[i] -> ', asesores[i])
            celda = hoja.cell(row=row_number+2+y, column=column_number+4)
            celda.value = asesores[i]

    else:
        asesores = info['Asesor(es)']['nombre']
        celda.value = asesores
    
    # print(asesores)
    

    # Variable Num Economico
    celda = hoja.cell(row=row_number+2, column=column_number+5)

    celda.alignment = alignment_header_tabla
    celda.font = font_tabla_info
    celda.border = border_header_tabla
    celda.value = 'N/A'

    # Blank space
    celda = hoja.cell(row=row_number+2, column=column_number+6)
    celda.border = border_header_tabla

    celda = hoja.cell(row=row_number+2, column=column_number+8)
    celda.border = border_header_tabla

    # MARTES, JUEVES y VIERNES
    celda = hoja.cell(row=row_number+2, column=column_number+7)

    celda.alignment = alignment_header_tabla
    celda.font = font_tabla_info
    celda.border = border_header_tabla
    celda.value = horario_del_curso

    celda = hoja.cell(row=row_number+2, column=column_number+9)

    celda.alignment = alignment_header_tabla
    celda.font = font_tabla_info
    celda.border = border_header_tabla
    celda.value = horario_del_curso

    celda = hoja.cell(row=row_number+2, column=column_number+10)

    celda.alignment = alignment_header_tabla
    celda.font = font_tabla_info
    celda.border = border_header_tabla
    celda.value = horario_del_curso


def crear_coeficiente(row_number, column_number, hoja, data):
    print('generando coeficiente...')

    header = ['PROFESOR', 'No. Económico', 'UEA', 'CLAVE', 'GRUPO', 'COEFICIENTE DE PARTICIPACIÓN', 'HORAS PARCIALES', 'HORAS TOTALES POR UEA']
    fill_header = PatternFill(start_color='a8d08d', end_color='a8d08d', fill_type='solid')
    font_header = Font(name='Arial', size=10, bold=True, italic=False, color='000000')
    alignment_header = Alignment(horizontal='center', vertical='center')
    border_header = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    grupopt1 = []
    grupopt2 = []
    grupopt3 = []

    grupos = [grupopt1, grupopt2, grupopt3]



    ### imprimiendo la cabecera de la tabla
    for index, i in enumerate(header):
        celda = hoja.cell(row=row_number, column=column_number+index, value=i)
        celda.fill = fill_header
        celda.font = font_header
        celda.alignment = alignment_header
        celda.border = border_header

    ### llenando la información de los grupos 
    for grupo in data:
        if grupo.get('PT') == '1':
            grupopt1.append(grupo)
        elif grupo.get('PT') == '2':
            grupopt2.append(grupo)
        else: 
            grupopt3.append(grupo)

    ### imprimir data
    contador_posicion = 1

    for index, grupo in enumerate(grupos):
        for i in grupo:
            
            asesores = i['Asesor(es)']
            cantidad_alumnos = len(i['Alumno(s)'])
            horas_parciales = cantidad_alumnos*2
            horas_por_alumno = 2

            if i['PT'] == '1':
                pt='I'
                clave = '450218'
                grupo = 'DJ01T'
                horas_totales_por_UEA = 9
            elif i['PT'] == '2':
                pt='II'
                clave = '450219'
                grupo = 'DK01T'
                horas_totales_por_UEA = 10

            else:
                pt='III'
                clave = '450220'
                grupo = 'DL01T'
                horas_totales_por_UEA = 10

            coeficiente_por_hora = horas_por_alumno / horas_totales_por_UEA

            if (coeficiente_por_hora * cantidad_alumnos > 1):
                coeficiente_de_participacion = 1 * len(asesores)
            else:
                coeficiente_de_participacion = round(coeficiente_por_hora * cantidad_alumnos, 2)


            if (isinstance(asesores, list)):
                # print('list')
                suma_coeficiente_de_participacion = len(asesores) * coeficiente_de_participacion
                suma_horas_parciales = len(asesores) * horas_parciales

                for index, asesor in enumerate(asesores):
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number, value=asesor['nombre'])
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+1, value=asesor['numero_economico'])
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+2, value= 'PROYECTO TERMINAL '+ pt)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+3, value= clave)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+4, value= grupo)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+5, value= coeficiente_de_participacion)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+6, value= horas_parciales)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+7, value= horas_totales_por_UEA)
                contador_posicion = contador_posicion + index

            else:
                # print('not a list')

                suma_coeficiente_de_participacion = coeficiente_de_participacion
                suma_horas_parciales = horas_parciales

                celda = hoja.cell(row=row_number+contador_posicion, column=column_number, value=asesores['nombre'])
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+1, value=asesores['numero_economico'])
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+2, value= 'PROYECTO TERMINAL '+ pt)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+3, value= clave)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+4, value= grupo)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+5, value= coeficiente_de_participacion)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+6, value= horas_parciales)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+7, value= horas_totales_por_UEA)


                ### SUMA
            celda = hoja.cell(row=row_number+contador_posicion+1, column=column_number, value= 'SUMA')
            celda = hoja.cell(row=row_number+contador_posicion+1, column=column_number+5, value= suma_coeficiente_de_participacion)
            celda = hoja.cell(row=row_number+contador_posicion+1, column=column_number+6, value= suma_horas_parciales)


            contador_posicion = celda.row + 1
#---------------------------------------------------------------
#---------------------------------------------------------------
### funcion que recibe el archivo externo, se leen los datos del 
# .txt subido desde flask

def bajarArchivo(archivoGlo1):
    try:
        # Abre el archivo en modo lectura con codificación utf-8
        with open(archivoGlo1, 'r', encoding='utf-8') as file:
            # Lee el contenido del archivo
            contenido = file.read()
            archivo = contenido
        
        # Retorna el contenido del archivo
        print(f"Archivo '{archivoGlo1}' leído exitosamente.")
        print(f"Archivo '{archivo}' a ocupar.")
        return archivo

    except FileNotFoundError:
        print(f"Error: El archivo '{archivoGlo1}' no fue encontrado.")
    except IOError:
        print(f"Error: No se pudo leer el archivo '{archivoGlo1}'.")
    except Exception as e:
        print(f"Ocurrió un error inesperado: {e}")

### funcion principal, donde se leen los datos del 
# .txt
def main(archivoGlobal1):
    
    with open(directorio_archiGruposDePtExcell + archivoGlobal1,'r', encoding="utf-8") as file:
    #with open(directorio_archiGruposDePtExcell + archivo,'r', encoding="utf-8") as file:
        print(f"Archivo que se usara en el main:´{archivoGlobal1}")
        print(f"Ruta que se usara en el main:´{directorio_archiGruposDePtExcell}")
      
        
        # Inicializa variables temporales para almacenar información de un proyecto
        titulo = ""
        asesores = ""
        alumno = ""
        matricula = ""
        pt = ""

        
        # Itera sobre cada línea del archivo
        for line in file:
            # Elimina los espacios en blanco al principio y al final de la línea
            line = line.strip()
            print(f"Line:´{line}")

     

            # Verifica si la línea contiene información relevante
            if line.startswith('Título:'):
                titulo = line[len('Título:'):].strip()

            elif line.startswith('Asesor(es):'):
                # Verifica si la línea contiene una coma para dividir los nombres de los revisores
                if ',' in line:
                    asesores = [rev.strip() for rev in line[len('Asesor(es):'):].split(',')]
                    asesores_lista = []
                    for rev in asesores:
                        indice = rev.index('.')
                        titulado = indice - 1

                        if rev[titulado] == 'a':
                            genero = 'F'
                        else:
                            genero = 'M'

                        numero_economico_asesor = get_numero_economico(manejo_de_nombre(rev))
                        asesores_lista.append({'nombre': rev, 'genero': genero, 'numero_economico': numero_economico_asesor})

                    asesores = asesores_lista
                else:
                    asesor = line[len('Asesor(es):'):].strip()

                    indice = asesor.index('.')
                    titulado = indice - 1

                    if asesor[titulado] == 'a':
                        genero = 'F'
                    else:
                        genero = 'M'
                        
                    
                    # print('rev -> ',asesor)
                    # print('nombre -> ', (manejo_de_nombre(asesor)))
                    # print('asesor -> ',get_numero_economico(manejo_de_nombre(asesor)))
                    numero_economico_asesor = get_numero_economico(manejo_de_nombre(asesor))
                    # if ( numero_economico_asesor == None): numero_economico_asesor = 'No se ha encontrado'
                    asesor = {'nombre': asesor, 'genero': genero, 'numero_economico': numero_economico_asesor}
                    asesores = asesor

            elif line.startswith('Alumno:'):
                alumno = line[len('Alumno:'):].strip()

            elif line.startswith('Matrícula:'):
                matricula = line[len('Matrícula:'):].strip()
                
            elif line.startswith('PT:'):
                pt = line[len('PT:'):].strip()
                
                grupo = {
                        'Título': titulo,
                        'Asesor(es)': asesores,
                        'Alumno': alumno,
                        'Matrícula': matricula,
                        'PT':pt
                    }
                
                ### creando los grupos
                manejo_grupo(grupo)

    ### debuggeando
    # print(grupos)

    ### verificar si la lista se ha llenado
        if grupos:
         print('generando excell con base en los grupos generados...')
        # print(grupos)
        # workbook= openpyxl.load_workbook(directorio_ruta+plantilla_excell)

        ### hoja de excell que necesitamos
        nombre_de_hoja = 'GruposPT'


        ### 1. copiar la info del template a una nueva hoja
        
        # craendo hoja con nombre de hoja y el numero del trimestre
        nueva_hoja = workbook.create_sheet(title=nombre_de_hoja+'-'+trimestre)

        # ### variables globales de estilo
        font_proyecto_terminal = Font(name='Arial', size=10, bold=True, italic=False, color='000000')
        fill_proyecto_terminal = PatternFill(start_color='ffe598', end_color='ffe598', fill_type='solid')
        fill_cambios_solicitados = PatternFill(start_color='ffc000', end_color='ffc000', fill_type='solid')
        fill_proyecto_terminalII = PatternFill(start_color='c5e0b3', end_color='c5e0b3', fill_type='solid')
        fill_proyecto_terminalIII = PatternFill(start_color='99ffff', end_color='99ffff', fill_type='solid')



        # columna A expandida
        nueva_hoja.column_dimensions['A'].width = 34
        nueva_hoja.column_dimensions['B'].width = 12
        nueva_hoja.column_dimensions['E'].width = 20

        # primera celda
        nueva_hoja.cell(row=2, column=1, value='REGISTRO ACTUAL PROYECTO TERMINAL I')
        celda = nueva_hoja['A2']
        celda.font = font_proyecto_terminal
        celda.fill = fill_proyecto_terminal

        # segunda celda

        
# con base en el siguietne formato:
# {
#   'Asesor(es)': {'nombre': '', 'genero': ''} | [{'nombre': '', 'genero': ''},{'nombre': '', 'genero': ''}], 
#   'Alumno(s)': [{'nombre' : '', 'Matrícula': ''},{'nombre' : '', 'Matrícula': ''}] , 
#   'PT': '1' | '2' | '3'
# }
        

        info_tabla = {
            'Asesor(es)': 'POR ASIGNAR',
            'Alumno(s)': '',
            'PT': '1',
            'CLAVE': 450218,
            'GRUPO': 'DJ01T'
        }

        crear_tabla(3, 1, nueva_hoja, info_tabla)
        
        # Imprimiendo los grupos de PT I 
        celda = nueva_hoja['A7']
        celda.font = font_proyecto_terminal
        celda.fill = fill_cambios_solicitados
        celda.value = 'CAMBIOS SOLICITADOS'

        grupopt1 = []
        grupopt2 = [{
            'Asesor(es)': 'POR ASIGNAR',
            'Alumno(s)': '',
            'PT': '2',
            'CLAVE': 450219,
            'GRUPO': 'DK01T'
        }]
        
        grupopt3 = [{
            'Asesor(es)': 'POR ASIGNAR',
            'Alumno(s)': '',
            'PT': '3',
            'CLAVE': 450220,
            'GRUPO': 'DL01T'
        }]


        for diccionario in grupos:
            if diccionario.get('PT') == '1':
                grupopt1.append(diccionario)
            elif diccionario.get('PT') == '2':
                grupopt2.append(diccionario)
            elif diccionario.get('PT') == '3':
                grupopt3.append(diccionario)

        
        # print(grupopt1)

        for i in range(len(grupopt1)):

                    y = i * 12

                    info_tabla = {
                    'Asesor(es)': grupopt1[i]['Asesor(es)'],
                    'Alumno(s)': grupopt1[i]['Alumno(s)'],
                    'PT': grupopt1[i]['PT'],
                    'CLAVE': 450218,
                    'GRUPO': 'DJ01T'
                    }
                    crear_tabla(9+y, 1, nueva_hoja, info_tabla)
                    marcador = y

        marcador_continua = marcador+20
        celda = nueva_hoja.cell(row=marcador_continua, column=1, value='REGISTRO ACTUAL PROYECTO TERMINAL II') 
        celda.font = font_proyecto_terminal
        celda.fill = fill_proyecto_terminalII

        ### enviando la info de pt II
        for i in range(len(grupopt2)):
            y = i * 12
            
            info_tabla = {
                    'Asesor(es)': grupopt2[i]['Asesor(es)'],
                    'Alumno(s)': grupopt2[i]['Alumno(s)'],
                    'PT': grupopt2[i]['PT'],
                    'CLAVE': 450219,
                    'GRUPO': 'DK01T'
                    }
            marcador_pt_ii = y
            
            crear_tabla(marcador_continua+2+y, 1, nueva_hoja, info_tabla)

        ### enviando info de pt III
        marcador_continua_pt_ii = marcador_pt_ii+38+marcador
        # print('marcador_continua_pt_ii ->', marcador_continua_pt_ii)
        celda = nueva_hoja.cell(row=marcador_continua_pt_ii, column=1, value='REGISTRO ACTUAL PROYECTO TERMINAL III') 
        celda.font = font_proyecto_terminal
        celda.fill = fill_proyecto_terminalIII

        for i in range(len(grupopt3)):
            y = i * 12
            
            info_tabla = {
                    'Asesor(es)': grupopt3[i]['Asesor(es)'],
                    'Alumno(s)': grupopt3[i]['Alumno(s)'],
                    'PT': grupopt3[i]['PT'],
                    'CLAVE': 450219,
                    'GRUPO': 'DK01T'
                    }
            
            crear_tabla(marcador_continua_pt_ii+2+y, 1, nueva_hoja, info_tabla)

        
        ### 3. generación de Coeficiente
        nombre_de_hoja = 'Coeficientes'
        nueva_hoja = workbook.create_sheet(title=nombre_de_hoja+'-'+trimestre)
        
        ### creando la tabla con base en la información
        crear_coeficiente(1,1,nueva_hoja,grupos)
        
        nueva_hoja.column_dimensions['A'].width = 34
        nueva_hoja.column_dimensions['B'].width = 15
        nueva_hoja.column_dimensions['C'].width = 34
        nueva_hoja.column_dimensions['D'].width = 15
        nueva_hoja.column_dimensions['E'].width = 15
        nueva_hoja.column_dimensions['F'].width = 35
        nueva_hoja.column_dimensions['G'].width = 30
        nueva_hoja.column_dimensions['H'].width = 30

        nueva_hoja.row_dimensions[1].height = 30


        ### 4. guardar los cambios en el libro de trabajo
        workbook.save(directorio_archiGeneradosGruposDePtExcell + excell_resultado)

        generar_txt_estudiantes_PT()
    

#else:
#print('Archivo con formato incorrecto!')


if __name__ == '__main__':

     # Define la ruta del archivo que deseas leer
    #ruta_archivo = 'C:/Users/serle/Desktop/PROYECTO/utilTech/GruposDePTExcell/archivosSubidos/asignacionAsesoresPT_26-01-24_20241127_202718.txt'  # Cambia esto a la ruta de tu archivo
    #archivo = bajarArchivo(ruta_archivo)
    #main(archivo)
    '''
    nombre_de_archivos = [f for f in os.listdir(directorio_ruta) if os.path.isfile(os.path.join(directorio_ruta, f))]

    while True:
        try:
            # Imprimir opciones
            print('Escribe el nombre de archivo de entrada (.txt)')
            for archivo in nombre_de_archivos:
                print('- ' + archivo)

            # Verificar si la respuesta está vacía o tiene espacios en blanco
            if not archivoGlobal1:
                os.system('cls' if os.name == 'nt' else 'clear')
                print('El nombre no puede estar vacío. Vuelve a intentarlo.')
                continue
            elif ' ' in archivoGlobal1:
                os.system('cls' if os.name == 'nt' else 'clear')
                print('El nombre tiene un espacio en blanco. Vuelve a intentarlo.')
                continue

            # Verificar si el nombre existe en la lista de archivos
            if archivoGlobal1 in nombre_de_archivos:
                main(archivoGlobal1)
                break
            else:
                os.system('cls' if os.name == 'nt' else 'clear')
                print('El archivo no existe en el directorio. Vuelve a intentarlo.')

        except Exception as e:
            os.system('cls' if os.name == 'nt' else 'clear')
            print(f"Opción inválida: {e}")
    '''

'''
 while True:
    nombre_de_archivos = [f for f in os.listdir(directorio_ruta) if os.path.isfile(os.path.join(directorio_ruta, f))]
    if archivoGlobal1 in nombre_de_archivos:
                main(archivoGlobal1)
                break
                
    else:
        os.system('cls' if os.name == 'nt' else 'clear')
        print('El archivo no existe en el directorio. Vuelve a intentarlo.')
'''