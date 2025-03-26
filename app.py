from flask import Flask, jsonify, request, render_template
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from werkzeug.utils import secure_filename
from datetime import date
from openpyxl import load_workbook
from configConstants import directorio_oficiosCargaAcademica, path_resultado, directorio_archivos_generados, directorio_archivos_subidos, archivosGenerados1_folder, directorio_archiGeneradosGruposDePtExcell, plantilla1, form_data, directorio_automatico

import openpyxl 
import os
import pandas as pd
import openpyxl
import secrets


app = Flask(__name__)


#######################################################################################
#######################################################################################
#######################################################################################
#Claves
app.secret_key = 'UtilTech' #Se establece la clave secreta manualmente con la cadena
app.secret_key = secrets.token_hex(16) #Se sobrescribe la clave anterior con un valor aleatorio generado

#######################################################################################
#######################################################################################
#######################################################################################
#Seccion generadora del trimestre lectivo
TRIMESTRE_FILE = "./GruposDePTExcell/Plantilla/trimestre_actual.txt"

def obtener_trimestre():
    hoy = date.today()
    anio = hoy.year % 100  # Obtener los últimos dos dígitos del año
    
    trimestres = [
        (date(hoy.year, 10, 21), date(hoy.year + 1, 1, 24), 'O'),
        (date(hoy.year, 2, 10), date(hoy.year, 5, 9), 'I'),
        (date(hoy.year, 5, 26), date(hoy.year, 8, 15), 'P')
    ]
    
    for inicio, fin, letra in trimestres:
        if inicio <= hoy <= fin:
            return f"{anio}-{letra}"
    
    # Si no está en ninguno de los rangos, significa que estamos fuera del periodo lectivo
    return f"{anio}-P" if hoy > trimestres[2][1] else f"{anio - 1}-O"
#Seccion donde se mantiene el trimestre fijo en la generacion de los otros documentos
def obtener_trimestre_fijo():
    # Si el archivo ya existe, usamos el trimestre almacenado
    if os.path.exists(TRIMESTRE_FILE):
        with open(TRIMESTRE_FILE, "r") as f:
            return f.read().strip()
    # Si no existe, generamos el trimestre y lo guardamos
    with open(TRIMESTRE_FILE, "w") as f:
        f.write(trimestre)
    return trimestre

#######################################################################################
#######################################################################################
#######################################################################################
#DEFINICION DE LAS VARIABLES GLOBALES
archivo_txt = None
workbook = None
grupos = []
trimestre = obtener_trimestre_fijo() #'24-O' #Trimestre lectivo / cambio cada trimestre de acuerdo a la fecha actual
profesores = []

path = directorio_oficiosCargaAcademica + "\\Plantilla\\propuesta.xlsx" # Esta si es variable

# Asegurarse de que las carpetas existen
for form in form_data.values():
    os.makedirs(form['subidos'], exist_ok=True)
    os.makedirs(form['generados'], exist_ok=True)

#######################################################################################
#######################################################################################
#######################################################################################
#SECCION DE LA FUNCION QUE CARGA EL INDEX
@app.route('/', methods=['GET','POST'])
def index():
    # Get form_id from query on the url
    form_id = request.args.get('form_id', 'formulario1')
    print("Este es el ID del formulario a ejecutarse: ", form_id)
    return render_template('index.html', 
                           list_Archivos=listaArchivos(form_id, "subidos"),)
                           #file_path=uploaded_file_path)

#######################################################################################
#######################################################################################
#######################################################################################
#SECCION DE LA FUNCION QUE SE ENCRAGA DE LA SUBIDA DE LOS ARCHIVOS
@app.route('/upload', methods=['POST'])
def upload_file():
    #Implementacion de las variables globales
    global form_data

    form_id = request.form.get('form_id')
    #Impresion de la seleccion del usuario en el form a ejecutarse de archivos a subirse
    print("Este es el ID del formulario a ejecutarse: ", form_id)

    if form_id in form_data:
        file_key = form_data[form_id]['file_key']
        folder = form_data[form_id]['subidos']
        file = request.files[file_key]

        #PARTE DEL FORMATO DEL NOMBRE QUE TOMARA EL ARCHIVO SUBIDO CON EL RENOMBRE
        timestamp =datetime.now().strftime('%Y%m%d_%H%M%S')

        filename = secure_filename(file.filename)
        name, ext = os.path.splitext(filename) #SEPARA EL NOMBRE DE SU TIPO DE EXTENSION
        new_filename = f"{name}_{timestamp}{ext}" #SE PROCEDE A RENOMBRAR EL ARCHIVO SUBIDO POR EL USUARIO
        upload_path = os.path.join(directorio_automatico, folder, new_filename) 
        file.save(upload_path)

        #Impresion de los archivos subidos
        print(f"Archivo subido correctamente: {file}")
        print(f"El archivo se guardó en {upload_path}")
        print(f"Archivo con time: {new_filename}")

        return render_template('index.html',
                                #FOLDERS DONDE SE ALMACENARON LOS ARCHIVOS SUBIDOS
                                archivosGenerados_folder=folder,
                                #SE PASA EL FORM SELECCIONADO
                                form_id=form_id,
                                #FUNCION PARA LA VISUALIZACION DE LOS ARCHIVOS EN LA CARPETA
                                list_Archivos = listaArchivos(form_id, 'subidos'),
                                #NOMBRES DE LOS ARCHIVOS SUBIDOS
                                archivo1 = file.filename,
                                #MENSAJES DONDE SE ALOJO EL ARCHIVO SUBIDO
                                message = f"Y este se guardo en:\n{upload_path}",
                                #SE DEJA VER AL USUARIO EL DIV CORRESPONDIENTE A LA RUTA DONDE SE GUARDO
                                div9_block = True
                               )

#######################################################################################
#######################################################################################
#######################################################################################
#SECCION DE LA FUNCION QUE SE ENCRAGA DE LA SUBIDA DE LOS ARCHIVOS
def listaArchivos(form_id, tipo):
    global form_data
    print(f"Formulario: {form_id}, Tipo: {tipo}")
    print(f"Formulario: {form_data}")
    if form_id in form_data and tipo in form_data[form_id]:
        path = form_data[form_id][tipo]
        # Si el tipo es un diccionario, se listan los archivos en cada subdirectorio
        if isinstance(path, dict):
            print(f"Archivo SUBIDO / GENERADO con time y ruta de directorio: {path}")
            return {key: os.listdir(sub_path) for key, sub_path in path.items()}
        
        # Si es una ruta simple, se lista directamente
        return os.listdir(path)
    
    return []  # Retorna una lista vacía si el form_id o tipo no son válidos

#######################################################################################
#######################################################################################
#######################################################################################
#FUNCION QUE PERMITE ELEMINAR LOS ARCHIVOS SUBIDOS POR EL USUARIO
@app.route('/deleteSub/<filename>', methods=['GET'])
def delete_file_subidos(filename):
    file_path = os.path.join(directorio_archivos_subidos, filename)

    # Aquí debes definir los valores para form_id y tipo
    form_id = 'formulario1'  # Cambia esto según tu lógica
    if os.path.isfile(file_path):
        try:
            os.remove(file_path)
            return render_template('index.html',
                               list_Archivos = listaArchivos(form_id, "subidos"),
                               show_subir = True)
        except Exception as e:
            return jsonify({'error': f'No se pudo eliminar el archivo: {str(e)}'}), 500
    else:
        return jsonify({'error': 'Archivo no encontrado'}), 404
    
#######################################################################################
#######################################################################################
#######################################################################################
#FUNCION QUE PERMITE ELEMINAR LOS ARCHIVOS GENERADOS POR EL USUARIO
@app.route('/deleteGen/<filename>', methods=['POST'])
def delete_file_generados(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER1'], filename)

    # Aquí debes definir los valores para form_id y tipo
    form_id = 'formulario1'  # Cambia esto según tu lógica
    if os.path.isfile(file_path):
        try:
            os.remove(file_path)
            return render_template('index.html',
                               list_Archivos = listaArchivos(form_id, "generados"),
                               show_subir = True)
        except Exception as e:
            return jsonify({'error': f'No se pudo eliminar el archivo: {str(e)}'}), 500
    else:
        return jsonify({'error': 'Archivo no encontrado'}), 404
    

#######################################################################################
#######################################################################################
#######################################################################################
#FUNCIONAMIENTO DEL PROGRAMA GENERADOR DE PLANTILLA PROGRAMA 1
#######################################################################################
#######################################################################################
#######################################################################################


#FUNCION QUE TRANSFORMA EL ARCHIVO CSV SUBIDO A ARCHIVO TXT CON LA ESTRUCTURA
def procesar_archivo_csv(archivo_para_procesar):
    global archivo_txt
    # Definir el nombre del archivo TXT que se generará
    archivo_txt = f"{archivo_para_procesar.split('.')[0]}.txt".replace('ArchivosSubidos', 'ArchivosGenerados')
    try:
        # Ejecutar la función para la generación del archivo
        convierte_csv_a_txt(archivo_para_procesar)
        # Retorna la ruta completa del archivo generado TXT
    except Exception as e:
        print(f"Error al procesar el archivo CSV: {e}")
    return archivo_txt

#######################################################################################
#######################################################################################
#FUNCION ENCARGADA DE CONVERTIR EL ARCHIVO CSV PARA EL CORRECTO FUNCIONAMIENTO DEL PROGRAMA 1
def convierte_csv_a_txt(archivo_csv):
    global archivo_txt
    try:
        # Lee el archivo CSV
        print(f"Leyendo el archivo CSV: {archivo_csv}")
        df = pd.read_csv(archivo_csv)
        print("Archivo CSV leído correctamente")
    except Exception as e:
        raise ValueError(f"Error al leer el archivo CSV: {e}")

    try:
        # Leer el archivo TXT para escritura
        print(f"Abriendo el archivo TXT para escritura: {archivo_txt}")
        with open(archivo_txt, 'w', encoding='utf-8') as f:
            for index, row in df.iterrows():
                try:
                    # Obtener el Titulo del proyecto terminal
                    if row['TituloProyecto']:
                        titulo = row['TituloProyecto']
                    else:
                        raise ValueError("El título del proyecto no está disponible")
                    print(f"Título del proyecto: {titulo}")

                    # Obtener los asesores
                    asesores = []
                    if pd.notna(row['ApellidosDelAsesor1']) and pd.notna(row['NombresDelAsesor1']):
                        asesores.append(f"{row['ApellidosDelAsesor1']} {row['NombresDelAsesor1']}")
                    if pd.notna(row['ApellidosDelAsesor2']) and pd.notna(row['NombresDelAsesor2']):
                        asesores.append(f"{row['ApellidosDelAsesor2']} {row['NombresDelAsesor2']}")
                    if pd.notna(row['ApellidosDelAsesor3']) and pd.notna(row['NombresDelAsesor3']):
                        asesores.append(f"{row['ApellidosDelAsesor3']} {row['NombresDelAsesor3']}")
                    print(f"Asesores: {asesores}")

                    # Obtener los revisores
                    revisores = []
                    if pd.notna(row['ApellidosDelRevisor1']) and pd.notna(row['NombresDelRevisor1']):
                        revisores.append(f"{row['ApellidosDelRevisor1']} ({row['NombresDelRevisor1']})")
                    if pd.notna(row['ApellidosDelRevisor2']) and pd.notna(row['NombresDelRevisor2']):
                        revisores.append(f"{row['ApellidosDelRevisor2']} ({row['NombresDelRevisor2']})")
                    print(f"Revisores: {revisores}")

                    # Obtener alumno y matricula
                    matricula_nombre = row['MatriculayNombre']
                    if not matricula_nombre:
                        raise ValueError("La matrícula y el nombre del alumno no están disponibles")
                    print(f"Matrícula y nombre: {matricula_nombre}")

                    # Separación del nombre y la matrícula
                    if ' - ' in matricula_nombre:
                        matricula, alumno = matricula_nombre.split(' - ')
                    else:
                        parts = matricula_nombre.split(' ')
                        matricula = parts[0]
                        alumno = ' '.join(parts[1:])
                    print(f"Alumno: {alumno}, Matrícula: {matricula}")

                    # Eliminación del valor PT repetitivo
                    pt = row['PT']
                    if not pt:
                        raise ValueError("El valor de PT no está disponible")
                    pt = pt.replace("PT", "").strip()
                    print(f"PT: {pt}")

                    # Escribir la estructura de cada alumno de proyecto terminal
                    f.write(f"Titulo: {titulo}\n")
                    f.write(f"Asesor(es): {', '.join(asesores)}\n")
                    f.write(f"Revisor(es): {', '.join(revisores)}\n")
                    f.write(f"Alumno: {alumno}\n")
                    f.write(f"Matrícula: {matricula}\n")
                    f.write(f"PT: {pt}\n\n")
                    print(f"Fila {index} procesada correctamente")
                except Exception as e:
                    raise ValueError(f"Error al procesar la fila {index}: {e}")
        print("Archivo TXT generado correctamente")
    except Exception as e:
        raise ValueError(f"Error al escribir el archivo TXT: {e}")

#######################################################################################
#######################################################################################
#FUNCION QUE SE ENCARGA DE COLOCAR EL NUMERO ECONOMICO A LOS PROFESORES
def get_numero_economico(nombre_del_profe, codigoGen1):
    try:
        if len(profesores) == 0:
            configProfesores()
        # Busca el número económico del profesor
        numero_economico = None
        for profe in profesores:
            if clean_name(nombre_del_profe) in clean_name(profe['nombre']):
                numero_economico = profe['numero_economico']
        if numero_economico is None:
            raise ValueError(f"Número económico no encontrado para el profesor {nombre_del_profe}")
        return numero_economico
    except Exception as e:
        raise RuntimeError(f"Error al obtener el número económico: {e}")
    
#######################################################################################
#######################################################################################
def configProfesores():
    #--------------------------------------------------------------
    #HOJA DE LOS PROFESORES
    hoja_profesores = openpyxl.load_workbook(plantilla1)['Profesores']
    #Recorrer la hoja de profesores
    for cell in hoja_profesores.iter_rows(min_row=3, max_row=hoja_profesores.max_row, min_col=1, max_col=2, values_only=True):
        if cell[0] is not None and cell[1] is not None:
            profesores.append({'nombre': cell[0], 'numero_economico': cell[1]})
        else:
            break

#######################################################################################
#######################################################################################
# Función para pasar los nombres a minusculas, sin espacios y sin acentos
def clean_name(name):
    name = name.lower()
    name = name.replace(' ', '')
    name = name.replace('á', 'a')
    name = name.replace('é', 'e')
    name = name.replace('í', 'i')
    name = name.replace('ó', 'o')
    name = name.replace('ú', 'u')
    return name

#######################################################################################
#######################################################################################
#FUNCION QUE MANEJA EL NOMBRE DE LOS PROFESORES EN CASO DE NO TENER SEGUNDO NOMBRE
def manejo_de_nombre(nombre):
    try:
        if '.' in nombre:
            nombre = nombre.split(' ', 1)[-1]
        else:
            pass
        nombre_temporal = nombre.split(' ')  # Separa la estructura del nombre
        # Si el ultimo elemento es igual a SIN, se quita
        if nombre_temporal[-1] == 'SIN':
            nombre_temporal.pop(-1)
        # Juntando las secciones del nombre
        nombre_final = ' '.join(nombre_temporal)
        # Regresa el nombre ya procesado
        return nombre_final
    except Exception as e:
        print(f"Error al manejar el nombre: {e}")
        return nombre  # Retorna el nombre original en caso de error
    
#######################################################################################
#######################################################################################
#FUNCION PARA GENERAR EL DOCUMENTO TXT QUE CONTENDRA EL NOMBRE DE LOS ESTUDIANTES Y A QUE PT PERTENECEN
def generar_txt_estudiantes_PT():
    print('Generando el archivo Estudiantes de PT')
    try:
        global archivo_txt
        # Quitar extensión a archivo_txt
        archivo_txt_sin_extension = archivo_txt.split('.')[0]
        print(f"Archivo TXT sin extensión: {archivo_txt_sin_extension}")
        # Crear el archivo TXT
        with open(f"{archivo_txt_sin_extension}_Estudiantes_PT.txt", "w", encoding="UTF-8") as file:
            grupos_de_PT = ['1', '2', '3']
            # Loop para agregar la información
            for element in grupos_de_PT:
                file.write('\nPT' + element + '\n')
                for grupo in grupos:
                    if grupo['PT'] == element:
                        for alumno in grupo['Alumno(s)']:
                            file.write(manejo_de_nombre(alumno['nombre']) + ' ' + alumno['matrícula'] + '\n')
        print("Archivo Estudiantes de PT generado correctamente")
    except Exception as e:
        print(f"Error al generar el archivo Estudiantes de PT: {e}")

#######################################################################################
#######################################################################################
#FUNCION PARA EL MANEJO DE LOS GRUPOS
def manejo_grupo(proyecto : dict):
    #Agrega la informacion del proyecto actual a LA LISTA DE GRUPOS
    grupo = {
        'Asesor(es)':[],
        'Alumno(s)':[],
        'PT':''
    }
    print('Manejando el grupo')
    print('grupos:', grupos)
    print('proyecto:', proyecto)
    #Revisa si la lista de grupos ESTA VACIA
    if(len(grupos) == 0):
        config_grupos_from_0(proyecto, grupo)
    else:
        # Verificando si hay algún elemento en 'grupos' con 'Asesor(es)' y 'PT' iguales a los del proyecto
        indice_coincidencia = next((indice for indice, grupo in enumerate(grupos) if grupo['Asesor(es)'] == proyecto['Asesor(es)'] and grupo['PT'] == proyecto['PT']), None)
        print('Indice de coincidencia:', indice_coincidencia)
        if indice_coincidencia is not None:
            ### agregando el alumno al mismo grupo de PT
            grupo_coincidente = grupos[indice_coincidencia]
            grupo_coincidente['Alumno(s)'].append({'nombre' : proyecto['Alumno'], 'matrícula': proyecto['Matricula']})
            print('grupo_coincidente:', grupo_coincidente)
        else:
            #Si no son del mismo grupo de PT
            config_grupos_from_0(proyecto, grupo)

def config_grupos_from_0(proyecto, grupo):
    grupo['Asesor(es)'] = proyecto['Asesor(es)']
    grupo['Alumno(s)'].append( {'nombre' : proyecto['Alumno'], 'matrícula': proyecto['Matricula']} )
    grupo['PT'] = proyecto['PT']
    grupos.append(grupo)

#######################################################################################
#######################################################################################
#FUNCION PARA CREAR LA TABLA EN EL ARCHIVO XLSX
def crear_tabla(row_number, column_number, hoja, info):
    #Estilos globales de la tabla a crear
    alignment_header_tabla = Alignment(horizontal='center', vertical='center')
    font_header_tabla = Font(name='Arial', size=8, bold=True, italic=False, color='000000')
    font_header_info = Font(name='Arial', size=7, color='000000')
    font_tabla_info = Font(name='Arial', size=11, color='000000')
    border_header_tabla = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    horario_del_curso='14-17'

    # Nombre del curso
    hoja.merge_cells(start_row=row_number, start_column=column_number, end_row=(row_number+1), end_column=column_number)
    celda = hoja.cell(row=row_number, column=column_number)


    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'NOMBRE DEL CURSO'

        # Clave
    hoja.merge_cells(start_row=row_number, start_column=(column_number+1), end_row=(row_number+1), end_column=(column_number+1))
    celda = hoja.cell(row=row_number, column=column_number+1)

    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'CLAVE'


    # Grupo
    hoja.merge_cells(start_row=row_number, start_column=(column_number+2), end_row=(row_number+1), end_column=(column_number+2))
    celda = hoja.cell(row=row_number, column=column_number+2)


    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'GRUPO'


    # Cupo 
    celda = hoja.cell(row=row_number, column=column_number+3)

    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'CUPO'

    # Max 
    celda = hoja.cell(row=row_number+1, column=column_number+3)

    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'MAX.'


    # Profesor
    hoja.merge_cells(start_row=row_number, start_column=(column_number+4), end_row=(row_number+1), end_column=(column_number+4))
    celda = hoja.cell(row=row_number, column=column_number+4)
    
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'PROFESOR'


    # No. eco
    hoja.merge_cells(start_row=row_number, start_column=(column_number+5), end_row=(row_number+1), end_column=(column_number+5))

    celda = hoja.cell(row=row_number, column=column_number+5)
    
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    celda.value = 'No. ECON.'


    # Horario y Aula
    hoja.merge_cells(start_row=row_number, start_column=column_number+6, end_row=(row_number), end_column=column_number+10)
    hoja.cell(row=row_number, column=(column_number+6), value='HORARIO  Y  AULA')
    celda = hoja.cell(row=row_number, column=(column_number+6))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla
    hoja.cell(row=row_number, column=(column_number+10)).border = border_header_tabla

    # LUNES, MARTES, MIERCOLES, JUEVES, VIERNES
    hoja.cell(row=row_number+1, column=(column_number+6), value='LUNES')
    celda = hoja.cell(row=row_number+1, column=(column_number+6))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla

    hoja.cell(row=row_number+1, column=(column_number+7), value='MARTES')
    celda = hoja.cell(row=row_number+1, column=(column_number+7))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla

    hoja.cell(row=row_number+1, column=(column_number+8), value='MIÉRCOLES')
    celda = hoja.cell(row=row_number+1, column=(column_number+8))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla


    hoja.cell(row=row_number+1, column=(column_number+9), value='JUEVES')
    celda = hoja.cell(row=row_number+1, column=(column_number+9))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla

    hoja.cell(row=row_number+1, column=(column_number+10), value='VIERNES')
    celda = hoja.cell(row=row_number+1, column=(column_number+10))
    celda.alignment = alignment_header_tabla
    celda.font = font_header_tabla
    celda.border = border_header_tabla



    if info['PT'] == '1': 
        pt = 'I' 
        horas_por_UEA = 9
    elif info['PT'] == '2': 
        pt = 'II'
        horas_por_UEA = 10
    else: 
        pt = 'III'
        horas_por_UEA = 10

    ### tabla observaciones
    if (info['Asesor(es)'] != 'POR ASIGNAR'):

        hoja.cell(row=row_number+1, column=(column_number+12), value='OBSERVACIONES')
        celda = hoja.cell(row=row_number+1, column=(column_number+12))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        hoja.cell(row=row_number+1, column=(column_number+13), value='Horas por alumno')
        celda = hoja.cell(row=row_number+1, column=(column_number+13))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla
        
        hoja.cell(row=row_number+1, column=(column_number+14), value='Alumnos')
        celda = hoja.cell(row=row_number+1, column=(column_number+14))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla
        
        hoja.cell(row=row_number+1, column=(column_number+15), value='Coeficiente por hr')
        celda = hoja.cell(row=row_number+1, column=(column_number+15))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        hoja.cell(row=row_number+1, column=(column_number+16), value='Coeficiente de participación')
        celda = hoja.cell(row=row_number+1, column=(column_number+16))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        hoja.cell(row=row_number+1, column=(column_number+17), value='Horas parciales')
        celda = hoja.cell(row=row_number+1, column=(column_number+17))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        hoja.cell(row=row_number+1, column=(column_number+18), value='Horas por UEA')
        celda = hoja.cell(row=row_number+1, column=(column_number+18))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        ### info llenado
        hoja.cell(row=row_number+2, column=(column_number+12), value='Asignación de profesor e inscripción de alumnos')
        celda = hoja.cell(row=row_number+2, column=(column_number+12))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        ### horas por alumnos valor
        hoja.cell(row=row_number+2, column=(column_number+13), value=2)
        celda = hoja.cell(row=row_number+2, column=(column_number+13))
        celda_horas_por_alumno = celda.value
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        ### Alumnos valor
        cantidad_de_alumnos = len(info['Alumno(s)']) 

        hoja.cell(row=row_number+2, column=(column_number+14), value=cantidad_de_alumnos)
        celda = hoja.cell(row=row_number+2, column=(column_number+14))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla
        
        ### Coeficiente por hr
        ### valor = horas por alumno / horas por UEA
        hoja.cell(row=row_number+2, column=(column_number+15), value=( round(celda_horas_por_alumno / horas_por_UEA, 2)))
        celda = hoja.cell(row=row_number+2, column=(column_number+15))
        celda_coeficiente_por_hora = celda.value
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla


        ### Coeficiente de part
        hoja.cell(row=row_number+2, column=(column_number+16), value= cantidad_de_alumnos* celda_coeficiente_por_hora)
        celda = hoja.cell(row=row_number+2, column=(column_number+16))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        ### Horas parciales
        hoja.cell(row=row_number+2, column=(column_number+17), value= cantidad_de_alumnos* 2)
        celda = hoja.cell(row=row_number+2, column=(column_number+17))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla


        ### Horas por UEA
        hoja.cell(row=row_number+2, column=(column_number+18), value=horas_por_UEA)
        celda = hoja.cell(row=row_number+2, column=(column_number+18))
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla
        celda.border = border_header_tabla

        ### tabla nombre de los alumnos:
        ### DEJAR INSCRITOS A LOS SIGUIENTES ALUMNOS:
        celda = hoja.cell(row=row_number+4, column=column_number)
        celda.value = 'DEJAR INSCRITOS A LOS SIGUIENTES ALUMNOS:'
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla

        celda = hoja.cell(row=row_number+6, column=column_number)
        celda.value = 'Nombre'
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla

        celda = hoja.cell(row=row_number+6, column=column_number+1)
        celda.value = 'Matricula'
        celda.alignment = alignment_header_tabla
        celda.font = font_header_tabla

        for index, alumno in enumerate(info['Alumno(s)']):
            celda_nombre = hoja.cell(row=row_number + 7 + index, column=column_number)
            celda_nombre.value = manejo_de_nombre(alumno['nombre'])

            celda_matricula = hoja.cell(row=row_number + 7 + index, column=column_number + 1)
            celda_matricula.value = alumno['matrícula']


    ### Llenando info
    # PROYECTO TERMINAL I
    celda = hoja.cell(row=row_number+2, column=column_number)


    celda.alignment = alignment_header_tabla
    celda.font = font_header_info
    celda.border = border_header_tabla
    celda.value = 'PROYECTO TERMINAL '+ pt

    # print('info ->', info)
    
    # Variable clave
    celda = hoja.cell(row=row_number+2, column=column_number+1)

    celda.alignment = alignment_header_tabla
    celda.font = font_header_info
    celda.border = border_header_tabla
    celda.value = int(info['CLAVE'])


    # Variable Grupo
    celda = hoja.cell(row=row_number+2, column=column_number+2)

    celda.alignment = alignment_header_tabla
    celda.font = Font(name='Calibri', size=11, color='000000')
    celda.border = border_header_tabla
    celda.value = info['GRUPO']

    

    # Variable Cupo máx
    celda = hoja.cell(row=row_number+2, column=column_number+3)

    celda.alignment = alignment_header_tabla
    celda.font = Font(name='Calibri', size=11, color='000000')
    celda.border = border_header_tabla
    celda.value = 30

    
    # Variable Profesor
    celda = hoja.cell(row=row_number+2, column=column_number+4)

    celda.alignment = alignment_header_tabla
    celda.font = Font(name='Calibri', size=11, color='000000')
    celda.border = border_header_tabla

    if info['Asesor(es)'] == 'POR ASIGNAR': 
        asesores = 'POR ASIGNAR'
        celda.value = asesores
    elif  isinstance( info['Asesor(es)'], list):
        # print('Its a list of professors')
        asesores = []

        for professor in info['Asesor(es)']:
            asesores.append(professor['nombre'])

        ### agregando el primer profe
        # print(asesores)
        celda.value = asesores[0]
        asesores.pop(0)

        for i in range(len(asesores)):
            y = i +1
            # print('valor de y -> ', y)
            # print('asesores[i] -> ', asesores[i])
            celda = hoja.cell(row=row_number+2+y, column=column_number+4)
            celda.value = asesores[i]

    else:
        asesores = info['Asesor(es)']['nombre']
        celda.value = asesores
    
    # print(asesores)
    

    # Variable Num Economico
    celda = hoja.cell(row=row_number+2, column=column_number+5)

    celda.alignment = alignment_header_tabla
    celda.font = font_tabla_info
    celda.border = border_header_tabla
    celda.value = 'N/A'

    # Blank space
    celda = hoja.cell(row=row_number+2, column=column_number+6)
    celda.border = border_header_tabla

    celda = hoja.cell(row=row_number+2, column=column_number+8)
    celda.border = border_header_tabla

    # MARTES, JUEVES y VIERNES
    celda = hoja.cell(row=row_number+2, column=column_number+7)

    celda.alignment = alignment_header_tabla
    celda.font = font_tabla_info
    celda.border = border_header_tabla
    celda.value = horario_del_curso

    celda = hoja.cell(row=row_number+2, column=column_number+9)

    celda.alignment = alignment_header_tabla
    celda.font = font_tabla_info
    celda.border = border_header_tabla
    celda.value = horario_del_curso

    celda = hoja.cell(row=row_number+2, column=column_number+10)

    celda.alignment = alignment_header_tabla
    celda.font = font_tabla_info
    celda.border = border_header_tabla
    celda.value = horario_del_curso

#######################################################################################
#######################################################################################
#FUNCION ENCARGADA EN LA GENERACION DE COEFICIENTES
def crear_coeficientes(row_number, column_number, hoja, data):
    print('Generando los coeficientes de los asesores y sus alumnos')

    #CABECERA A CREARCE EN LA SECCION DE COEFICIENTES
    header = ['PROFESOR', 'No. Económico', 'UEA', 'CLAVE', 'GRUPO', 'COEFICIENTE DE PARTICIPACIÓN', 'HORAS PARCIALES', 'HORAS TOTALES POR UEA']
    fill_header = PatternFill(start_color='a8d08d', end_color='a8d08d', fill_type='solid')
    font_header = Font(name='Arial', size=10, bold=True, italic=False, color='000000')
    alignment_header = Alignment(horizontal='center', vertical='center')
    border_header = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    #creacion de arreglos de los PTS
    grupopt1 = []
    grupopt2 = []
    grupopt3 = []

    #creacion de conjunto de todos los arreglos
    grupos = [grupopt1, grupopt2, grupopt3]

    #Impresion de de la cabecera generada en la tabla 
    for index, i in enumerate(header):
        celda = hoja.cell(row=row_number, column=column_number+index, value=i)
        celda.fill = fill_header
        celda.font = font_header
        celda.alignment = alignment_header
        celda.border = border_header

    #Llenando la informacion de los GRUPOS
    for grupo in data:
        if grupo.get('PT') == '1':
            grupopt1.append(grupo)
        elif grupo.get('PT') == '2':
            grupopt2.append(grupo)
        else: 
            grupopt3.append(grupo)
        
    #Imprimir el data
    contador_posicion = 1

    for index, grupo in enumerate(grupos):
        for i in grupo:
            #Colocacion de los asesores / alumnos 
            asesores = i['Asesor(es)']
            cantidad_alumnos = len(i['Alumno(s)'])
            horas_parciales = cantidad_alumnos*2
            horas_por_alumno = 2
            #Colocaion de la clave / grupo de cada PT
            if i['PT'] == '1':
                pt='I'
                clave = '450218'
                grupo = 'DJ01T'
                horas_totales_por_UEA = 9
            elif i['PT'] == '2':
                pt='II'
                clave = '450219'
                grupo = 'DK01T'
                horas_totales_por_UEA = 10

            else:
                pt='III'
                clave = '450220'
                grupo = 'DL01T'
                horas_totales_por_UEA = 10
            #CALCULANDO LOS COEFICIENTES POR HORA
            coeficiente_por_hora = horas_por_alumno / horas_totales_por_UEA #Se divide las horas cada alumno y las horas totales por UEA
            
            if(coeficiente_por_hora * cantidad_alumnos > 1):
                coeficiente_de_participacion = 1 * len(asesores)
            else:
                coeficiente_de_participacion = round(coeficiente_por_hora * cantidad_alumnos, 2)

            if (isinstance(asesores, list)):
                #Lista
                suma_coeficiente_de_participacion = len(asesores) * coeficiente_de_participacion
                suma_horas_parciales = len(asesores) * horas_parciales

                for index, asesor in enumerate(asesores):
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number, value=asesor['nombre'])
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+1, value=asesor['numero_economico'])
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+2, value= 'PROYECTO TERMINAL '+ pt)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+3, value= clave)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+4, value= grupo)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+5, value= coeficiente_de_participacion)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+6, value= horas_parciales)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+7, value= horas_totales_por_UEA)
                contador_posicion = contador_posicion + index

            else:
                suma_coeficiente_de_participacion = coeficiente_de_participacion
                suma_horas_parciales = horas_parciales

                celda = hoja.cell(row=row_number+contador_posicion, column=column_number, value=asesores['nombre'])
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+1, value=asesores['numero_economico'])
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+2, value= 'PROYECTO TERMINAL '+ pt)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+3, value= clave)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+4, value= grupo)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+5, value= coeficiente_de_participacion)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+6, value= horas_parciales)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+7, value= horas_totales_por_UEA)


            #SUMA
            celda = hoja.cell(row=row_number+contador_posicion+1, column=column_number, value= 'SUMA')
            celda = hoja.cell(row=row_number+contador_posicion+1, column=column_number+5, value= suma_coeficiente_de_participacion)
            celda = hoja.cell(row=row_number+contador_posicion+1, column=column_number+6, value= suma_horas_parciales)


            contador_posicion = celda.row + 1

#######################################################################################
#######################################################################################
#FUNCION MAIN PARA GRUPOS DE PT
def main(codigoGen1):
    with open(archivosGenerados1_folder + codigoGen1,'r',encoding="utf-8") as file:
         print(f"El archivo que se usara en el main de GRUPOS DE PT:'{codigoGen1}")
         #Se incializan las variables temporales para almacenar la informacion del proyecto
         titulo = ""
         asesores = ""
         matricula = ""
         pt = ""

         #Itera sobre cada linea del archivo
         for line in file:
            # Elimina los espacios en blanco al principio y al final de la línea
            line = line.strip()
            # Verifica si la línea contiene la informacion de TITULO
            if line.startswith('Título:'):
                titulo = line[len('Título:'):].strip()
            elif line.startswith('Asesor(es):'):
                asesores__to_get_numero_economico = []
                #Verifica si la linea que se esta leyendo  contiene coma para dividir los nombres de los revisores
                if ',' in line:
                    asesores_lista_in_line = list(set([rev.strip() for rev in line[len('Asesor(es):'):].split(',')]))
                    print("asesores_lista_in_line", asesores_lista_in_line)
                    for rev in asesores_lista_in_line:
                        asesores__to_get_numero_economico.append(rev)
                else:
                    asesor = line[len('Asesor(es):'):].strip()
                    print("asesor", asesor)
                    asesores__to_get_numero_economico.append(asesor)
                asesores = []
                for asesor in asesores__to_get_numero_economico:
                    try:
                        genero = get_genre(asesor)
                        numero_economico_asesor = get_numero_economico(manejo_de_nombre(asesor), codigoGen1)
                        asesor = {'nombre': asesor, 'genero': genero, 'numero_economico': numero_economico_asesor}
                        asesores.append(asesor)
                    except Exception as e:
                        print(f"Error al obtener el número económico del asesor: {e}")
                        continue

            elif line.startswith('Alumno:'):
                alumno = line[len('Alumno:'):].strip()

            elif line.startswith('Matrícula:'):
                matricula = line[len('Matrícula:'):].strip()
                
            elif line.startswith('PT:'):
                pt = line[len('PT:'):].strip()
                
                #Estructura a tomar de los datos recopilados
                grupo = {
                        'Título': titulo,
                        'Asesor(es)': asesores,
                        'Alumno': alumno,
                        'Matricula': matricula,
                        'PT':pt
                    }
                manejo_grupo(grupo) #Creando los grupos
         print('Grupos:', grupos)
         #Verificacion de la lista llenada
         if grupos:
            print('Generando excell con base en los grupos generados')
            nombre_de_hoja = 'GruposPt'
            #1.- Se copia la informacion de template a una nueva hoja
            #Se crea la hoja con el nombre y el numero de trimestre
            nueva_hoja = workbook.create_sheet(title=nombre_de_hoja+'-'+trimestre)

            #Se definen las variables del estilo
            font_proyecto_terminal = Font(name='Arial', size=10, bold=True, italic=False, color='000000')
            fill_proyecto_terminal = PatternFill(start_color='ffe598', end_color='ffe598', fill_type='solid')
            fill_cambios_solicitados = PatternFill(start_color='ffc000', end_color='ffc000', fill_type='solid')
            fill_proyecto_terminalII = PatternFill(start_color='c5e0b3', end_color='c5e0b3', fill_type='solid')
            fill_proyecto_terminalIII = PatternFill(start_color='99ffff', end_color='99ffff', fill_type='solid')

            #Columna A expandida
            nueva_hoja.column_dimensions['A'].width = 34
            nueva_hoja.column_dimensions['B'].width = 12
            nueva_hoja.column_dimensions['E'].width = 20

            #Primera Celda
            nueva_hoja.cell(row=2, column=1, value='REGISTRO ACTUAL PROYECTO TERMINAL I')
            celda = nueva_hoja['A2']
            celda.font = font_proyecto_terminal
            celda.fill = fill_proyecto_terminal

            #Creacion de la informacion de la tabla para los PT1, PT2 Y PT3
            info_tabla = {
            'Asesor(es)': 'POR ASIGNAR',
            'Alumno(s)': '',
            'PT': '1',
            'CLAVE': 450218,
            'GRUPO': 'DJ01T'
             }
             #Grupos de PT 1
            crear_tabla(3, 1, nueva_hoja, info_tabla)
        
            celda = nueva_hoja['A7']
            celda.font = font_proyecto_terminal
            celda.fill = fill_cambios_solicitados
            celda.value = 'CAMBIOS SOLICITADOS'

            grupopt1 = []
            grupopt2 = [{
            'Asesor(es)': 'POR ASIGNAR',
            'Alumno(s)': '',
            'PT': '2',
            'CLAVE': 450219,
            'GRUPO': 'DK01T'
            }]
        
            grupopt3 = [{
            'Asesor(es)': 'POR ASIGNAR',
            'Alumno(s)': '',
            'PT': '3',
            'CLAVE': 450220,
            'GRUPO': 'DL01T'
            }]

    #Incroporacion de los grupos en el diccionario
    for diccionario in grupos:
        if diccionario.get('PT') == '1':
            grupopt1.append(diccionario)
        elif diccionario.get('PT') == '2':
            grupopt2.append(diccionario)
        elif diccionario.get('PT') == '3':
            grupopt3.append(diccionario)
        marcador = 0

        for i in range(len(grupopt1)):
                y = i * 12
                info_tabla = {
                    'Asesor(es)': grupopt1[i]['Asesor(es)'],
                    'Alumno(s)': grupopt1[i]['Alumno(s)'],
                    'PT': grupopt1[i]['PT'],
                    'CLAVE': 450218,
                    'GRUPO': 'DJ01T'   
                }
                crear_tabla(9+y, 1, nueva_hoja, info_tabla)
                marcador = y

    #Se envia la informacion de PT II 
    marcador_continua = marcador+20
    celda = nueva_hoja.cell(row=marcador_continua, column=1, value='REGISTRO ACTUAL PROYECTO TERMINAL II') 
    celda.font = font_proyecto_terminal
    celda.fill = fill_proyecto_terminalII

    #Se manda la informacion del PT II
    for i in range(len(grupopt2)):
                y = i * 12
                #Creacion de la tabla con los valores
                info_tabla = {
                    'Asesor(es)': grupopt2[i]['Asesor(es)'],
                    'Alumno(s)': grupopt2[i]['Alumno(s)'],
                    'PT': grupopt2[i]['PT'],
                    'CLAVE': 450219,
                    'GRUPO': 'DK01T'
                    }
                marcador_pt_ii = y
                #Creacion de la tabla
                crear_tabla(marcador_continua+2+y, 1, nueva_hoja, info_tabla)


    #Se envia la informacion de PT II
    marcador_continua_pt_ii = marcador_pt_ii+38+marcador
    celda = nueva_hoja.cell(row=marcador_continua_pt_ii, column=1, value='REGISTRO ACTUAL PROYECTO TERMINAL III')
    celda.font = font_proyecto_terminal
    celda.fill = fill_proyecto_terminalIII

    #Se manda la informacion del PT III
    for i in range(len(grupopt3)):
            y = i * 12
            
            info_tabla = {
                    'Asesor(es)': grupopt3[i]['Asesor(es)'],
                    'Alumno(s)': grupopt3[i]['Alumno(s)'],
                    'PT': grupopt3[i]['PT'],
                    'CLAVE': 450219,
                    'GRUPO': 'DK01T'
                    }
            crear_tabla(marcador_continua_pt_ii+2+y, 1, nueva_hoja, info_tabla)

    #3.- Generacion de coeficientes
    nombre_de_hoja = 'Coeficientes'
    nueva_hoja = workbook.create_sheet(title=nombre_de_hoja+'-'+trimestre)
   
    #Creando la tabla con base en la informacion
    crear_coeficiente(1,1,nueva_hoja,grupos)
        
    nueva_hoja.column_dimensions['A'].width = 34
    nueva_hoja.column_dimensions['B'].width = 15
    nueva_hoja.column_dimensions['C'].width = 34
    nueva_hoja.column_dimensions['D'].width = 15
    nueva_hoja.column_dimensions['E'].width = 15
    nueva_hoja.column_dimensions['F'].width = 35
    nueva_hoja.column_dimensions['G'].width = 30
    nueva_hoja.column_dimensions['H'].width = 30

    nueva_hoja.row_dimensions[1].height = 30

    #4.- Guardar los cambios en el libro de trabajo
    try:
        timestamp =datetime.now().strftime('%Y%m%d_%H%M%S')
        workbook_file_to_save = f"{directorio_archiGeneradosGruposDePtExcell}\\PlantillaAsignacionCoeficientesPT-{trimestre}_{timestamp}.xlsx"
        print(f"Guardando el archivo en: {workbook_file_to_save}")
        workbook.save(workbook_file_to_save)
        generar_txt_estudiantes_PT()
        print(f"Archivo generados con éxito en: {workbook_file_to_save}")
        return workbook_file_to_save
    except Exception as e:
        print(f"Error al guardar el archivo o generar el TXT: {e}")
        raise Exception(f"Error al guardar el archivo o generar el TXT: {e}")
    
#######################################################################################
#######################################################################################
def get_genre(name):
    indice = name.index(' ')
    titulado = indice - 1
    if name[titulado] == 'a':
        return 'F'
    else:
        return 'M'



#######################################################################################
#######################################################################################
#######################################################################################
#FUNCION GENERADORA DEL DOCUMENTO EXCEL Y EL TXT DEL PRIMER MODULO GRUPOS DE PT (ARCHIVO ESPECIFICO)
#######################################################################################
#######################################################################################
#######################################################################################
@app.route('/GruposPT/<filename>', methods=['GET'])
def generar_GruposPT(filename):
    global directorio_automatico
    form_id = request.form.get('form_id')
    fileGenerated = doGenerarGruposPT(filename, form_id)
    if fileGenerated.success:
            return render_template('generar.html',
                                            directorio_automatico2=directorio_automatico,
                                            archivosGenerados1_folder2=archivosGenerados1_folder,
                                            #list_Archivos2=listar_archivos(form_id, tipo),
                                            message=f"El archivo se guardó en {fileGenerated.txt_filename}",
                                            archivosGlobales=[fileGenerated.filename],
                                            codigosGen=[fileGenerated.codigoGen1],
                                            #generated_files=generated_files,
                                            list_Archivos=listar_archivos_subidos(),
                                            archivos=[fileGenerated.archivo],
                                            message2=f"El archivo: {fileGenerated.txt_filename}\n se generó con éxito en {directorio_archiGeneradosGruposDePtExcell}",
                                            txt_filenames=[fileGenerated.txt_filename],
                                            #excel_filename=excel_filename,
                                            div9_block=True)
    else:
        return render_template('index.html',
                                            directorio_automatico2=directorio_automatico,
                                            archivosGenerados1_folder2=archivosGenerados1_folder,
                                            message=f"Hubo una falla al procesar el archivo",
                                            archivosGlobales=[],
                                            codigosGen=[],
                                            list_Archivos=listar_archivos_subidos(),
                                            message2=f"Archivo CSV no encontrado",
                                            txt_filenames=[],
                                            div9_block=True)

####################################################################################### 
#######################################################################################
#######################################################################################
#FUNCION GENERADORA DEL DOCUMENTO EXCEL Y EL TXT DEL PRIMER MODULO GRUPOS DE PT (MULTI ARCHIVOS)
#######################################################################################
#######################################################################################
#######################################################################################
@app.route('/GruposPTMulti', methods=['POST'])
def generar_GruposPTMulti():
    global directorio_automatico
    form_id = request.form.get('form_id') or 'formulario1'
    print('request.form.items()',request.form.items())
    options = []
    for key, value in request.form.items():
        if key.endswith(('csv', 'xls', 'xlsx', 'txt')):  # Ensure this is a tuple of strings
            options.append(value)
    print('options:', options)  # Add a colon for better readability
    filesGenerated = []
    for option in options:
        filesGenerated.append(doGenerarGruposPT(option, form_id))
    print('filesGenerated',filesGenerated)
    successedFiles = []
    failedFiles = []
    for fileGenerated in filesGenerated:
        print('fileGenerated',fileGenerated)
        if fileGenerated.get("success") is not None and fileGenerated.get("success") is True:
            successedFiles.append(fileGenerated)
        else:
            failedFiles.append(fileGenerated)
    message = ''
    if successedFiles and len(successedFiles) > 0:
        message = f"Los archivos se guardaron en {', '.join([file.get('txt_filename') for file in successedFiles if file.get('txt_filename')])}"
    if failedFiles and len(failedFiles) > 0:
        message = message + f"Hubo una falla al procesar los siguientes archivos: {', '.join([file.get('filename') for file in failedFiles if file.get('filename')])}"
    message2 = ''
    if successedFiles and len(successedFiles) > 0:
        message2 = f"Los archivos {', '.join([file.get('txt_filename') for file in successedFiles if file.get('txt_filename')])} se generaron con éxito en {directorio_archiGeneradosGruposDePtExcell}"
    if failedFiles and len(failedFiles) > 0:
        message2 = message2 + f"Los archivos {', '.join([file.get('filename') for file in failedFiles if file.get('filename')])}, tuvieron el siguiente error: Archivo CSV no encontrado"
    archivosGlobales = [fileGenerated.get("filename") for fileGenerated in successedFiles]
    codigosGen = [fileGenerated.get("codigoGen1") for fileGenerated in successedFiles]
    archivos = [fileGenerated.get("archivo") for fileGenerated in successedFiles]
    txt_filenames = [fileGenerated.get("txt_filename") for fileGenerated in successedFiles]
    list_Archivos = listaArchivos(form_id, "subidos")
    list_Archivos_Generados = listaArchivos(form_id, "generados")
    print('message',message)
    print('message2',message2)
    print('successedFiles',successedFiles)
    print('failedFiles',failedFiles)
    print('directorio_automatico',directorio_automatico)
    print('archivosGenerados1_folder',archivosGenerados1_folder)
    print('list_Archivos',list_Archivos)
    print('archivosGlobales',archivosGlobales)
    print('codigosGen',codigosGen)
    print('archivos',archivos) #No trae campos
    print('txt_filenames',txt_filenames)
    print('list_Archivos_Generados',list_Archivos_Generados)
    return render_template('index.html', 
                            directorio_automatico2=directorio_automatico,
                            archivosGenerados1_folder2=archivosGenerados1_folder,
                            archivosGlobales=archivosGlobales,
                            codigosGen=codigosGen,
                            list_Archivos=list_Archivos,
                            archivos=archivos,
                            message=message,
                            message2=message2,
                            txt_filenames=txt_filenames,
                            list_Archivos_Generados = list_Archivos_Generados,
                            div9_block=True)


#######################################################################################
#######################################################################################
def doGenerarGruposPT(filename, form_id):
    global directorio_automatico

    folderSubidos = form_data[form_id]['subidos']
    folderGenerados = form_data[form_id]['generados']

    archivo_con_ruta = folderSubidos + filename
    print(f"filename que se usará en el main: {filename}")
    print(f"archivo_con_ruta que se usará en el main: {archivo_con_ruta}")  # RUTA COMPLETA DEL CSV SUBIDO POR EL USUARIO

    try:
        # Revisar que archivo_con_ruta incluya .csv
        hasDotCSV = archivo_con_ruta.endswith('.csv')
        if not hasDotCSV:
            raise ValueError("El archivo no tiene la extensión .csv")

        if os.path.exists(archivo_con_ruta):  # csv_path
            txt_filename = procesar_archivo_csv(archivo_con_ruta)  # csv_path

            # Asegúrate de que el archivo TXT existe antes de procesarlo
            if not os.path.exists(txt_filename):
                raise FileNotFoundError(f"El archivo TXT no se generó correctamente: {txt_filename}")

            print(f"El archivo txt tiene la ruta de: {txt_filename}")
            txt_name = os.path.basename(txt_filename)
            name, ext = os.path.splitext(txt_name)
            new_filename = f"{name}{ext}"
            codigoGen1 = new_filename
            print(f"El archivo txt tiene el nombre de: {codigoGen1}")

            # Llama a la función en main.py que procesará el archivo TXT
            if form_id == 'formulario1':
                coeficientes = main(codigoGen1)  # Suponiendo que 'main' es la función que deseas llamar en main.py
            elif form_id == 'formulario2':
                coeficientes = main2(codigoGen1)  # Suponiendo que 'main' es la función que deseas llamar en main.py
            print(f"El archivo coeficientes tiene el nombre de: {coeficientes}")

            # Renderizar la plantilla con los datos necesarios
            return {
                "txt_filename": txt_filename,
                "filename": filename,
                "codigoGen1": codigoGen1,
                "archivo": coeficientes,
                "success": True
            }
        else:
            raise FileNotFoundError(f"El archivo CSV no se encontró: {archivo_con_ruta}")

    except Exception as e:
        print(f"Error al generar los grupos PT: {e}")
        return {
            "filename": filename,
            "success": False
        }

#######################################################################################
#######################################################################################
#######################################################################################
#FUNCION QUE LISTA LOS ARCHIVOS
@app.route('/archivo/<filename>', methods=['GET'])
def view_file_generated(form_id, tipo):
    archivos_lista = []

    # Inicializar archivos_subidos_lista de manera segura
    archivos_subidos_lista = []

    # Verificar si los archivos han sido subidos (comprobando form_data)
    if form_id in form_data and tipo in form_data[form_id]:
        archivos_subidos_lista = [
            file for file in os.listdir(directorio_archivos_subidos)
            if file.endswith('.txt') or file.endswith('.xlsx')# and file != '~$PlantillaAsignacionCoeficientesPT-24-O.xlsx'
        ]
    
    # Listar archivos subidos
    archivos_lista.extend(archivos_subidos_lista)

    # Listar archivos generados
    archivos_generados_lista = [
        file for file in os.listdir(directorio_archivos_generados)
        if file.endswith('.txt') or file.endswith('.xlsx')# and file != '~$PlantillaAsignacionCoeficientesPT-24-O.xlsx'
    ]
    archivos_lista.extend(archivos_generados_lista)

    # Imprimir listas para depuración
    print(f"Archivos subidos: {archivos_subidos_lista}")
    print(f"Archivos generados: {archivos_generados_lista}")

    return archivos_lista

def listar_archivos_subidos():
    # Listar todos los archivos en la carpeta que sean .txt o .xlsx, excluyendo el archivo temporal
    archivos_lista = [
        file for file in os.listdir(directorio_archivos_subidos)
        if (file.endswith('.txt') or file.endswith('.xlsx')) #and file != '~$PlantillaAsignacionCoeficientesPT-24-O.xlsx'
    ]

    # Imprimir lista para depuración
    print(f"Archivos subidos: {archivos_lista}")

    return archivos_lista


def crear_coeficiente(row_number, column_number, hoja, data):
    print('generando coeficiente...')

    header = ['PROFESOR', 'No. Económico', 'UEA', 'CLAVE', 'GRUPO', 'COEFICIENTE DE PARTICIPACIÓN', 'HORAS PARCIALES', 'HORAS TOTALES POR UEA']
    fill_header = PatternFill(start_color='a8d08d', end_color='a8d08d', fill_type='solid')
    font_header = Font(name='Arial', size=10, bold=True, italic=False, color='000000')
    alignment_header = Alignment(horizontal='center', vertical='center')
    border_header = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))
    grupopt1 = []
    grupopt2 = []
    grupopt3 = []

    grupos = [grupopt1, grupopt2, grupopt3]



    ### imprimiendo la cabecera de la tabla
    for index, i in enumerate(header):
        celda = hoja.cell(row=row_number, column=column_number+index, value=i)
        celda.fill = fill_header
        celda.font = font_header
        celda.alignment = alignment_header
        celda.border = border_header

    ### llenando la información de los grupos 
    for grupo in data:
        if grupo.get('PT') == '1':
            grupopt1.append(grupo)
        elif grupo.get('PT') == '2':
            grupopt2.append(grupo)
        else: 
            grupopt3.append(grupo)

    ### imprimir data
    contador_posicion = 1

    for index, grupo in enumerate(grupos):
        for i in grupo:
            
            asesores = i['Asesor(es)']
            cantidad_alumnos = len(i['Alumno(s)'])
            horas_parciales = cantidad_alumnos*2
            horas_por_alumno = 2

            if i['PT'] == '1':
                pt='I'
                clave = '450218'
                grupo = 'DJ01T'
                horas_totales_por_UEA = 9
            elif i['PT'] == '2':
                pt='II'
                clave = '450219'
                grupo = 'DK01T'
                horas_totales_por_UEA = 10

            else:
                pt='III'
                clave = '450220'
                grupo = 'DL01T'
                horas_totales_por_UEA = 10

            coeficiente_por_hora = horas_por_alumno / horas_totales_por_UEA

            if (coeficiente_por_hora * cantidad_alumnos > 1):
                coeficiente_de_participacion = 1 * len(asesores)
            else:
                coeficiente_de_participacion = round(coeficiente_por_hora * cantidad_alumnos, 2)


            if (isinstance(asesores, list)):
                # print('list')
                suma_coeficiente_de_participacion = len(asesores) * coeficiente_de_participacion
                suma_horas_parciales = len(asesores) * horas_parciales

                for index, asesor in enumerate(asesores):
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number, value=asesor['nombre'])
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+1, value=asesor['numero_economico'])
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+2, value= 'PROYECTO TERMINAL '+ pt)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+3, value= clave)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+4, value= grupo)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+5, value= coeficiente_de_participacion)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+6, value= horas_parciales)
                    celda = hoja.cell(row=row_number+contador_posicion+index, column=column_number+7, value= horas_totales_por_UEA)
                contador_posicion = contador_posicion + index

            else:
                # print('not a list')

                suma_coeficiente_de_participacion = coeficiente_de_participacion
                suma_horas_parciales = horas_parciales

                celda = hoja.cell(row=row_number+contador_posicion, column=column_number, value=asesores['nombre'])
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+1, value=asesores['numero_economico'])
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+2, value= 'PROYECTO TERMINAL '+ pt)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+3, value= clave)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+4, value= grupo)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+5, value= coeficiente_de_participacion)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+6, value= horas_parciales)
                celda = hoja.cell(row=row_number+contador_posicion, column=column_number+7, value= horas_totales_por_UEA)


                ### SUMA
            celda = hoja.cell(row=row_number+contador_posicion+1, column=column_number, value= 'SUMA')
            celda = hoja.cell(row=row_number+contador_posicion+1, column=column_number+5, value= suma_coeficiente_de_participacion)
            celda = hoja.cell(row=row_number+contador_posicion+1, column=column_number+6, value= suma_horas_parciales)


            contador_posicion = celda.row + 1

#######################################################################################
#######################################################################################
#FUNCIONAMIENTO DEL PROGRAMA GENERADOR DE PLANTILLA PROGRAMA 2
#######################################################################################
#######################################################################################

#Abriendo el archivo de excel
#Este diccionario se llena de datos
diccionario_de_posiciones_de_tablas = []

### diccionario de datos de las tablas
diccionario_tablas = {
    'NOMBRE': [],
    'MATERIA': [],
    'GRUPO': [],
    'HORAS': []
}

#Guardando el documento
documento = pd.read_excel(path, sheet_name='FORMATO PROGRAMACIÓN', engine='openpyxl')
#######################################################################################
#######################################################################################
#Funcion para recorrer los valores
def recorrerValores(array, string):
    i=0
    while i < len(array[i]):
        for item in array[i] == string:
            if item == True:
                return i
        i += 1
    

#######################################################################################
#######################################################################################
### función para insertar datos al diccionario
def insertarDatosProfesAlDiccionario(diccionario, nombre, materia, grupo, horas):
    ### si los arreglos están vacíos, se agregar datos del primer profe a cada arreglo
        diccionario['NOMBRE'].append(nombre)
        diccionario['MATERIA'].append(materia)
        diccionario['GRUPO'].append(grupo)
        diccionario['HORAS'].append(horas)

#######################################################################################
#######################################################################################
def extraer_datos(archivo_entrada):
    
    path = directorio_automatico + archivo_entrada
    documento = pd.ExcelFile(path, engine='openpyxl')

    # Recorrer cada hoja en el archivo
    for nombre_hoja in documento.sheet_names:
        # Cargar la hoja actual
        df = pd.read_excel(path, sheet_name=nombre_hoja, engine='openpyxl')
        
        # Lista para almacenar las posiciones encontradas en la hoja actual
        posiciones = []

        # Recorrer cada celda en la hoja actual
        for i in range(len(df)):
            for j in range(len(df.columns)):
                if df.iloc[i, j] == 'NOMBRE DEL CURSO':
                    # Convertir índice de columna numérico a notación de letras
                    columna_letra = get_column_letter(j + 1)
                    # Almacenar la posición en formato Excel (por ejemplo, A1, B2)
                    posiciones.append(f"{columna_letra}{i + 1}")

        # Si se encontraron posiciones, agregarlas a la lista principal
        if posiciones:
            diccionario_de_posiciones_de_tablas.append({
                'pagina': nombre_hoja,
                'cabeceras': posiciones
            })

# Función para obtener el valor de una celda
def obtener_valor_celda(celda, nombre_hoja, archivo_entrada):
    path = directorio_automatico + archivo_entrada
    wb = load_workbook(filename=path)
    ws = wb[nombre_hoja]

    return ws[celda].value

#######################################################################################
#######################################################################################
def recorrer_tablas(data:dict, archivo_entrada):
    
    # documento_temporal = pd.read_excel(path, sheet_name= data['hoja'], engine='openpyxl')
    nombre_hoja = data.pop('hoja', None)

    #### recorrer los valores de la tabla y agregarlo a la info de arriba:
    # Recorrer el diccionario
    for indice, posiciones in data.items():
        # Obtener el número de la fila inicial
        fila_inicial = int(''.join(filter(str.isdigit, indice)))
        
        # Comenzar a partir de la fila inicial + 3
        fila_actual = fila_inicial + 3

        while True:
            # Construir la referencia de celda para cada columna
            celda_indice = f"{indice[0]}{fila_actual}"
            celda_profesor = f"{posiciones['PROFESOR'][0]}{fila_actual}"
            celda_grupo = f"{posiciones['GRUPO'][0]}{fila_actual}"
            celda_horas = f"{posiciones['HORAS'][0]}{fila_actual}"

            # Obtener los valores de las celdas
            valor_indice = obtener_valor_celda(celda_indice, nombre_hoja, archivo_entrada)
            valor_profesor = obtener_valor_celda(celda_profesor, nombre_hoja, archivo_entrada)
            valor_grupo = obtener_valor_celda(celda_grupo, nombre_hoja, archivo_entrada)
            valor_horas = obtener_valor_celda(celda_horas, nombre_hoja, archivo_entrada)

            # Verificar si el valor de la celda índice es nulo
            if valor_indice is None:
                break
            else:

                    if (len(diccionario_tablas['NOMBRE'] and diccionario_tablas['MATERIA'] and diccionario_tablas['HORAS']) < 0):
                    
                        ### si los arreglos están vacíos, se agregar datos del primer profe a cada arreglo
                        insertarDatosProfesAlDiccionario(diccionario_tablas, valor_profesor, valor_indice, valor_grupo, valor_horas)

                    else:
                        
                        ### si el profe existe en el diccionario, agregarlo
                        if(valor_profesor in diccionario_tablas['NOMBRE']):

                            ### sacar su posición
                            posicion = diccionario_tablas['NOMBRE'].index(valor_profesor)
                            ### agregamos los nuevos valores en la posición
                            diccionario_tablas['NOMBRE'].insert(posicion, valor_profesor)
                            diccionario_tablas['MATERIA'].insert(posicion, valor_indice)        
                            diccionario_tablas['GRUPO'].insert(posicion, valor_grupo)
                            diccionario_tablas['HORAS'].insert(posicion, valor_horas)
                        else:

                            ### agregamos los nuevos valores en la posición
                            insertarDatosProfesAlDiccionario(diccionario_tablas, valor_profesor, valor_indice, valor_grupo, valor_horas)

                    # Imprimir los valores
                    # print(f"Valor en {celda_indice}: {valor_indice}")
                    # print(f"Valor en {celda_profesor}: {valor_profesor}")
                    # print(f"Valor en {celda_grupo}: {valor_grupo}")
                    # print(f"Valor en {celda_horas}: {valor_horas}")
                    # print('---')

                    print(f"{valor_indice}")
                    print(f"{valor_profesor}")
                    print(f"{valor_grupo}")
                    print(f"{valor_horas}")
                    print('---')

            # Incrementar la fila actual
            fila_actual += 1

def vueltaDic(nombre, value):
    for item in value:
        if item['nombre'] == nombre:
            return True
        else:
            pass
df=pd.DataFrame(diccionario_tablas)

df['TOTAL DE HORAS'] = df.groupby('NOMBRE', sort=False)['HORAS'].transform('sum')
result = df.set_index(['NOMBRE','TOTAL DE HORAS','GRUPO'])

with pd.ExcelWriter(path, mode='a', if_sheet_exists='replace')as writer:
    result.to_excel(writer, sheet_name = 'RESULTADOS')

    workbook = load_workbook(path)
    worksheet = workbook['RESULTADOS']

    worksheet.column_dimensions['A'].width = 36
    worksheet.column_dimensions['B'].width = 17
    worksheet.column_dimensions['D'].width = 45

    workbook.save(path)
    workbook.close()

#######################################################################################
#######################################################################################
def recoleccion_indices(archivo_entrada):
    path = directorio_automatico + archivo_entrada

    for grupo in diccionario_de_posiciones_de_tablas:

        nombre_de_la_hoja = grupo['pagina']
        celdas = grupo['cabeceras']

        # Encabezados a buscar
        encabezados_a_buscar = ['PROFESOR', 'PROFESOR POSIBLE', 'GRUPO', 'HORAS']


        documento_temporal = pd.read_excel(path, sheet_name= nombre_de_la_hoja, engine='openpyxl')

        # Diccionario para almacenar los resultados
        celdas_encabezados_por_fila = {
            'hoja': nombre_de_la_hoja,
        }

        # Iterar sobre cada celda en la lista de celdas
        for celda in celdas:
            # Extraer la letra de la columna y el número de la fila
            columna_letra = ''.join([c for c in celda if c.isalpha()])  # Ej. 'B'
            fila_numero = int(''.join([c for c in celda if c.isdigit()]))  # Ej. 5

            # Obtener la fila específica
            fila_especifica = documento_temporal.iloc[fila_numero - 1]  # Restar 1 para índice basado en 0

            # Buscar los encabezados en la fila específica y construir la referencia de celda
            celdas_encabezados = {}
            for encabezado in encabezados_a_buscar:
                if encabezado in fila_especifica.values:
                    # Obtener la posición del encabezado
                    columna_indice = fila_especifica[fila_especifica == encabezado].index[0]
                    columna_numero = documento_temporal.columns.get_loc(columna_indice) + 1  # Índice basado en 1
                    columna_letra = get_column_letter(columna_numero)

                    # Construir la referencia de celda en formato Excel
                    celda_referencia = f"{columna_letra}{fila_numero}"

                    # Unificar 'PROFESOR' y 'PROFESOR POSIBLE' bajo la clave 'PROFESOR'
                    if encabezado in ['PROFESOR', 'PROFESOR POSIBLE']:
                        celdas_encabezados['PROFESOR'] = celda_referencia
                    else:
                        celdas_encabezados[encabezado] = celda_referencia

            # Almacenar las celdas encontradas para cada fila inicial
            celdas_encabezados_por_fila[celda] = celdas_encabezados


        recorrer_tablas(celdas_encabezados_por_fila, archivo_entrada)


    # Generación del archivo resultado
    if diccionario_tablas:
        df = pd.DataFrame(diccionario_tablas)
        df['TOTAL DE HORAS'] = df.groupby('NOMBRE', sort=False)['HORAS'].transform('sum')
        result = df.set_index(['NOMBRE', 'TOTAL DE HORAS', 'GRUPO'])

        # Crear un nuevo archivo Excel con una hoja llamada 'RESULTADOS'
        with pd.ExcelWriter(path_resultado, mode='w') as writer:  # Cambiamos a modo 'w' para crear un nuevo archivo
            result.to_excel(writer, sheet_name='RESULTADOS')

        # Cargar el workbook recién creado
        workbook = load_workbook(path_resultado)
        # Abrir la hoja del workbook
        worksheet = workbook['RESULTADOS']

        # Declarar el tamaño de las columnas
        worksheet.column_dimensions['A'].width = 36
        worksheet.column_dimensions['B'].width = 17
        worksheet.column_dimensions['D'].width = 45

        workbook.save(path_resultado)
        workbook.close()
    else:
        print('No se agregaron datos')

#######################################################################################
#######################################################################################
def main2(archivo_entrada:str):

    extraer_datos(archivo_entrada)

    if diccionario_de_posiciones_de_tablas:
        recoleccion_indices(archivo_entrada)
    else:
        print('no se ha encontrado ninguna tabla')
#######################################################################################
#######################################################################################
#######################################################################################




#######################################################################################
#######################################################################################
#######################################################################################
#FUNCION MAIN ENCARGADA DEL EXCELENTE FUNCIONAMIENTO
#######################################################################################
#######################################################################################
#######################################################################################
if __name__ == "__main__":
    app.run(debug=True)